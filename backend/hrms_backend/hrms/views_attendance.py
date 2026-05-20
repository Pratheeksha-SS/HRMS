import openpyxl
import re
import traceback
import random
import string
from datetime import timedelta, time, date as date_cls

from django.db import transaction
from django.db.models import Q, Sum, Avg
from django.utils import timezone

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from .models import Employee, EmployeeAttendance, AttendanceSummary, AttendanceUpload, User

# Accept multiple common date-header formats found in attendance excels.
# Examples: 01-Jan, 1-Jan, 01-Jan-2026, 01/Jan, 01 Jan
_DATE_RE = re.compile(r"^\s*\d{1,2}[-/\s]\s*[A-Za-z]{3}(?:[-/\s]\s*(\d{4}))?\s*$")



# ---------------------------------------------------------------------------
# XLS / XLSX loader
# ---------------------------------------------------------------------------

def _read_xls(file_obj):
    import xlrd
    file_obj.seek(0)
    wb = xlrd.open_workbook(file_contents=file_obj.read())
    ws = wb.sheet_by_index(0)
    rows = []
    for rx in range(ws.nrows):
        row = []
        for cx in range(ws.ncols):
            cell = ws.cell(rx, cx)
            if cell.ctype == 0:
                row.append(None)
            elif cell.ctype == 1:
                row.append(cell.value)
            elif cell.ctype == 2:
                v = cell.value
                if 0 < v < 1:
                    mins = round(v * 24 * 60)
                    h, m = divmod(mins, 60)
                    row.append("%02d:%02d" % (h, m))
                else:
                    row.append(int(v) if v == int(v) else v)
            elif cell.ctype == 3:
                dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                row.append(dt.strftime("%d-%b"))
            else:
                row.append(cell.value)
        rows.append(tuple(row))
    return rows


def _load_rows(file_obj):
    file_obj.seek(0)
    magic = file_obj.read(4)
    file_obj.seek(0)
    if magic[:4] == b"\xd0\xcf\x11\xe0":
        return _read_xls(file_obj)
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active
    return list(ws.iter_rows(values_only=True))


# ---------------------------------------------------------------------------
# Value parsers
# ---------------------------------------------------------------------------

def _parse_time(val):
    if not val:
        return None
    s = re.sub(r"\(SE\)", "", str(val)).strip()
    if s in ("", "00:00", "0:00"):
        return None
    parts = s.split(":")
    try:
        h, m = int(parts[0]), int(parts[1])
        return None if (h == 0 and m == 0) else time(h, m)
    except Exception:
        return None


def _parse_dur(val):
    if not val:
        return None
    s = str(val).strip()
    if s in ("", "00:00", "0:00"):
        return None
    parts = s.split(":")
    try:
        h, m = int(parts[0]), int(parts[1])
        return None if (h == 0 and m == 0) else timedelta(hours=h, minutes=m)
    except Exception:
        return None


def _parse_status(raw):
    if not raw:
        return "A", False, False, True, False, False
    s = str(raw).strip().upper()
    clean = s.replace(" ", "").replace("(", "").replace(")", "")
    is_mo = "MISSEDOUT" in clean
    is_wo = s.startswith("WO")
    is_p  = s.startswith("P") and not is_wo
    is_a  = s == "A"
    if is_wo:
        code = "WO"
    elif is_a:
        code = "A"
    elif is_p and is_mo:
        code = "P_MO"
    elif is_p:
        code = "P"
    else:
        code = "A"
    return code, is_wo, is_p, is_a, False, is_mo


def _month_num(abbr):
    return {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,
            "jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}.get(abbr.lower(), 1)


# ---------------------------------------------------------------------------
# Smart employee matching — by employee_code field only
# ---------------------------------------------------------------------------

def _find_emp_by_code(excel_code, excel_name=None):
    """
    Match an Excel employee_code to an Employee record.

    The Excel file contains codes like "Intern 44", "EMP001", "EXEC01".
    The admin must enter the exact same value in the Employee Code field
    when registering the employee.

    Priority:
      1. Employee.employee_code  — exact match (the dedicated field admin fills in)
      2. Employee.employee_id    — fallback for legacy records (e.g. EMPEX26001)
      3. Employee.user.username  — match by username
      4. Employee name           — match by exact or partial name
    """
    if not excel_code:
        return None, "not_registered"

    code = excel_code.strip()

    # 1. Match by the dedicated employee_code field (exact, case-insensitive)
    emp = Employee.objects.filter(employee_code__iexact=code).first()
    if emp:
        return emp, "matched"

    # 2. Fallback: match by employee_id (e.g. EMPEX26001)
    emp = Employee.objects.filter(employee_id__iexact=code).first()
    if emp:
        return emp, "matched"

    # 3. Match by username
    emp = Employee.objects.filter(user__username__iexact=code).first()
    if emp:
        return emp, "matched"

    # 4. Match by excel_name if provided
    if excel_name:
        name_clean = excel_name.strip()
        for e in Employee.objects.all():
            if e.get_full_name().strip().lower() == name_clean.lower():
                return e, "matched"
        parts = name_clean.split()
        if len(parts) >= 2:
            first, last = parts[0], parts[-1]
            emp = Employee.objects.filter(first_name__icontains=first, last_name__icontains=last).first()
            if emp:
                return emp, "matched"

    return None, "not_registered"


# ---------------------------------------------------------------------------
# Main parser
# ---------------------------------------------------------------------------

def parse_attendance_excel(file_obj, upload_obj):
    """
    Parse Elogixa Monthly Detailed Attendance Report (.xls or .xlsx).
    Returns (records_saved, records_failed, error_log_str, auto_created_list).
    """
    year  = upload_obj.year
    month = upload_obj.month

    try:
        rows = _load_rows(file_obj)
    except Exception as e:
        return 0, 0, "Could not read file: %s" % str(e), []

    # Find the date-header row
    date_row_idx = None
    date_cols    = {}

    for i, row in enumerate(rows):
        vals = [str(v).strip() if v else "" for v in row]
        hits = [(j, v) for j, v in enumerate(vals) if _DATE_RE.match(v)]
        if len(hits) >= 5:
            date_row_idx = i
            for j, v in hits:
                try:
                    # Parse formats like:
                    # 01-Jan | 1-Jan | 01/Jan | 01 Jan | 01-Jan-2026 (year part optional)
                    parts = re.split(r"[-/\s]+", v.strip())
                    # Expect: [day, month_abbr] or [day, month_abbr, optional_year]
                    day = int(parts[0])
                    month_part = parts[1]
                    date_cols[j] = date_cls(year, _month_num(month_part), day)

                except Exception:
                    pass
            break

    if date_row_idx is None or not date_cols:
        return 0, 0, (
            "Could not find the date header row. "
            "Make sure you are uploading the Elogixa Monthly Detailed Attendance Report "
            "and that the month/year match the file."
        ), []

    saved  = 0
    failed = 0
    errors = []
    auto_created = []   # list of dicts for newly created employees
    i = date_row_idx + 1
    n = len(rows)

    while i < n:
        row   = rows[i]
        vals  = [str(v).strip() if v is not None else "" for v in row]
        line  = " ".join(vals)

        if "Employee Code:-" not in line or "Employee Name:-" not in line:
            i += 1
            continue

        emp_code = emp_name = desig = ""
        want_code = want_name = want_desig = False
        for v in vals:
            if "Employee Code:-" in v:
                want_code = True; continue
            if want_code and v and not emp_code:
                emp_code = v; want_code = False
            if "Employee Name:-" in v:
                want_name = True; continue
            if want_name and v and not emp_name:
                emp_name = v; want_name = False
            if "Designation" in v and "-" in v:
                want_desig = True; continue
            if want_desig and v and not desig:
                desig = v; want_desig = False

        if not emp_code or not emp_name:
            i += 1
            continue

        in_row = out_row = late_row = early_row = ot_row = dur_row = st_row = None
        j = i + 1
        while j < min(i + 30, n):
            r    = rows[j]
            rv   = [str(v).strip() if v is not None else "" for v in r]
            first = next((v for v in rv if v), "")
            if   first == "In Time":   in_row   = rv
            elif first == "Out Time":  out_row  = rv
            elif first == "Late By":   late_row = rv
            elif first == "Early By":  early_row= rv
            elif first == "Total OT":  ot_row   = rv
            elif first == "Duration" and dur_row is None: dur_row = rv
            elif first == "Status":    st_row   = rv; break
            j += 1

        if st_row is None:
            i = j + 1
            continue

        # Match employee by code — no auto-creation
        matched, match_status = _find_emp_by_code(emp_code, emp_name)
        not_registered = (match_status == "not_registered")

        def _g(row, idx):
            return row[idx] if row and idx < len(row) else None

        try:
            with transaction.atomic():
                for col, rec_date in date_cols.items():
                    raw_st = _g(st_row, col)
                    code, is_wo, is_p, is_a, _, is_mo = _parse_status(raw_st)
                    late_v  = _g(late_row,  col)
                    early_v = _g(early_row, col)

                    EmployeeAttendance.objects.update_or_create(
                        employee_code=emp_code,
                        date=rec_date,
                        defaults=dict(
                            upload=upload_obj,
                            employee=matched,          # None if not registered
                            employee_name=emp_name,
                            designation=desig,
                            month=month,
                            year=year,
                            in_time=_parse_time(_g(in_row,  col)),
                            out_time=_parse_time(_g(out_row, col)),
                            late_by=_parse_dur(late_v),
                            early_by=_parse_dur(early_v),
                            overtime=_parse_dur(_g(ot_row,  col)),
                            duration=_parse_dur(_g(dur_row, col)),
                            status=code,
                            is_weekly_off=is_wo,
                            is_present=is_p,
                            is_absent=is_a,
                            is_late=bool(_parse_dur(late_v)),
                            is_early_exit=bool(_parse_dur(early_v)),
                            is_missed_out=is_mo,
                        )
                    )
                    saved += 1

            _build_summary(emp_code, emp_name, matched, month, year, upload_obj)

            if not_registered:
                auto_created.append({
                    "employee_code": emp_code,
                    "employee_name": emp_name,
                    "status": "not_registered",
                    "message": (
                        f"Employee Code '{emp_code}' not found in Employee module. "
                        f"Please register this employee and set their Employee Code to '{emp_code}'."
                    ),
                })

        except Exception as e:
            failed += 1
            errors.append("%s (%s): %s" % (emp_code, emp_name, str(e)[:200]))

        i = j + 1

    return saved, failed, "\n".join(errors), auto_created


def _build_summary(emp_code, emp_name, matched, month, year, upload_obj):
    recs = EmployeeAttendance.objects.filter(employee_code=emp_code, month=month, year=year)
    total_p  = recs.filter(is_present=True).count()
    total_a  = recs.filter(is_absent=True).count()
    total_wo = recs.filter(is_weekly_off=True).count()
    total_mo = recs.filter(is_missed_out=True).count()

    dur = late = early = ot = timedelta()
    for r in recs:
        if r.duration:  dur   += r.duration
        if r.late_by:   late  += r.late_by
        if r.early_by:  early += r.early_by
        if r.overtime:  ot    += r.overtime

    AttendanceSummary.objects.update_or_create(
        employee_code=emp_code, month=month, year=year,
        defaults=dict(
            employee_name=emp_name, employee=matched, upload=upload_obj,
            total_present=total_p, total_absent=total_a, total_leave=0,
            total_weekly_off=total_wo,
            total_working_days=total_p + total_a,
            total_duration_mins=int(dur.total_seconds()   // 60),
            total_late_mins=    int(late.total_seconds()  // 60),
            total_early_mins=   int(early.total_seconds() // 60),
            total_overtime_mins=int(ot.total_seconds()    // 60),
            missed_out_count=total_mo,
        )
    )


# ---------------------------------------------------------------------------
# Serialiser helpers
# ---------------------------------------------------------------------------

def _rec(r):
    return {
        "id": r.id, "date": str(r.date),
        "employee_code": r.employee_code, "employee_name": r.employee_name,
        "designation": r.designation,
        "in_time":  r.in_time.strftime("%H:%M")  if r.in_time  else None,
        "out_time": r.out_time.strftime("%H:%M") if r.out_time else None,
        "working_hours": r.working_hours,
        "late_by":  r.late_by_display, "early_by": r.early_by_display,
        "status": r.status,
        "status_display": dict(EmployeeAttendance.STATUS_CHOICES).get(r.status, r.status),
        "is_present": r.is_present, "is_absent": r.is_absent,
        "is_weekly_off": r.is_weekly_off, "is_late": r.is_late,
        "is_early_exit": r.is_early_exit, "is_missed_out": r.is_missed_out,
        "month": r.month, "year": r.year,
        # True if this attendance record is linked to a registered employee
        "is_registered": r.employee_id is not None,
        "employee_id_sys": r.employee.employee_id if r.employee else None,
        "department": r.employee.department if r.employee else "",
    }


def _summ(s):
    h, m = divmod(s.total_duration_mins, 60)
    return {
        "employee_code": s.employee_code, "employee_name": s.employee_name,
        "month": s.month, "year": s.year,
        "total_present": s.total_present, "total_absent": s.total_absent,
        "total_weekly_off": s.total_weekly_off, "total_working_days": s.total_working_days,
        "total_duration": "%02d:%02d" % (h, m),
        "total_late_mins": s.total_late_mins, "total_early_mins": s.total_early_mins,
        "total_overtime_mins": s.total_overtime_mins, "missed_out_count": s.missed_out_count,
        "attendance_percentage": s.attendance_percentage,
        "avg_working_hours": s.avg_working_hours,
    }


# ---------------------------------------------------------------------------
# API Views
# ---------------------------------------------------------------------------

class AttendanceUploadView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if request.user.role != "ADMIN":
            return Response({"error": "Admin only"}, status=403)

        file  = request.FILES.get("file")
        month = request.data.get("month")
        year  = request.data.get("year")

        if not file:
            return Response({"error": "No file uploaded. Please select an Excel file."}, status=400)
        if not month or not year:
            return Response({"error": "month and year are required."}, status=400)

        try:
            month, year = int(month), int(year)
        except (ValueError, TypeError):
            return Response({"error": "month and year must be integers."}, status=400)

        fname = file.name.lower()
        if not (fname.endswith(".xls") or fname.endswith(".xlsx")):
            return Response({"error": "Only .xls and .xlsx files are supported."}, status=400)

        upload = AttendanceUpload.objects.create(
            uploaded_by=request.user, file=file,
            original_filename=file.name, month=month, year=year, status="PROCESSING",
        )

        try:
            file.seek(0)
            saved, failed, err_log, auto_created = parse_attendance_excel(file, upload)
            upload.records_processed = saved
            upload.records_failed    = failed
            upload.error_log         = err_log or ""
            upload.status            = "DONE" if saved > 0 else "FAILED"
            upload.save()

            if saved == 0 and err_log:
                return Response({
                    "message": "Upload failed — no records were processed.",
                    "upload_id": upload.id,
                    "records_processed": 0,
                    "records_failed": failed,
                    "status": "FAILED",
                    "errors": err_log[:2000],
                    "auto_created_employees": [],
                }, status=422)

            return Response({
                "message": "Processed %d records, %d failed." % (saved, failed),
                "upload_id": upload.id,
                "records_processed": saved,
                "records_failed": failed,
                "status": upload.status,
                "errors": err_log[:1000] if err_log else None,
                # Employees from Excel whose code was not found in Employee module
                "unregistered_employees": [x for x in auto_created if x.get("status") == "not_registered"],
                "unregistered_count": sum(1 for x in auto_created if x.get("status") == "not_registered"),
                "auto_created_employees": auto_created,
                "auto_created_count": len(auto_created),
            })
        except Exception as e:
            traceback.print_exc()
            upload.status    = "FAILED"
            upload.error_log = str(e)
            upload.save()
            return Response({"error": "Server error during processing: %s" % str(e)}, status=500)

    def get(self, request):
        qs = AttendanceUpload.objects.all()[:20] if request.user.role == "ADMIN" \
             else AttendanceUpload.objects.filter(uploaded_by=request.user)[:10]
        return Response([{
            "id": u.id, "original_filename": u.original_filename,
            "month": u.month, "year": u.year,
            "records_processed": u.records_processed, "records_failed": u.records_failed,
            "status": u.status, "uploaded_at": u.uploaded_at.strftime("%d %b %Y %H:%M"),
        } for u in qs])


class AttendanceUploadDeleteView(APIView):
    """
    DELETE /attendance/upload/<id>/
    Admin-only. Removes the upload record, all linked EmployeeAttendance rows,
    all linked AttendanceSummary rows, and the physical file from disk.
    """
    permission_classes = [IsAuthenticated]

    def delete(self, request, pk):
        if request.user.role != "ADMIN":
            return Response({"error": "Admin only."}, status=403)

        try:
            upload = AttendanceUpload.objects.get(pk=pk)
        except AttendanceUpload.DoesNotExist:
            return Response({"error": "Upload not found."}, status=404)

        month = upload.month
        year  = upload.year
        fname = upload.original_filename

        try:
            with transaction.atomic():
                # 1. Count what we are about to delete (for the response message)
                att_count  = EmployeeAttendance.objects.filter(upload=upload).count()
                summ_count = AttendanceSummary.objects.filter(upload=upload).count()

                # 2. Delete daily attendance records linked to this upload
                EmployeeAttendance.objects.filter(upload=upload).delete()

                # 3. Delete monthly summaries linked to this upload
                AttendanceSummary.objects.filter(upload=upload).delete()

                # 4. Delete the physical file from disk (ignore if already gone)
                import os
                if upload.file and upload.file.name:
                    try:
                        file_path = upload.file.path
                        if os.path.isfile(file_path):
                            os.remove(file_path)
                    except Exception:
                        pass  # File already missing — not a blocker

                # 5. Delete the upload DB record itself
                upload.delete()

            return Response({
                "message": (
                    f"Upload '{fname}' ({month}/{year}) deleted successfully. "
                    f"Removed {att_count} attendance records and {summ_count} summaries."
                ),
                "deleted_attendance_records": att_count,
                "deleted_summaries": summ_count,
                "month": month,
                "year": year,
            }, status=200)

        except Exception as e:
            traceback.print_exc()
            return Response({"error": f"Failed to delete upload: {str(e)}"}, status=500)


class AttendanceListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user     = request.user
        month    = request.query_params.get("month")
        year     = request.query_params.get("year")
        emp_code = request.query_params.get("employee_code")
        emp_name = request.query_params.get("employee_name")
        dept     = request.query_params.get("department")

        qs = EmployeeAttendance.objects.select_related("employee").all()

        if user.role == "EMPLOYEE":
            try:
                emp = Employee.objects.get(user=user)
                qs = qs.filter(Q(employee=emp) | Q(employee_code=emp.employee_id))
            except Employee.DoesNotExist:
                return Response([])
        elif user.role == "MANAGER":
            dept_name = user.managed_department
            if not dept_name:
                return Response([])
            qs = qs.filter(
                Q(employee__department__iexact=dept_name) |
                Q(employee_code__in=Employee.objects.filter(
                    department__iexact=dept_name
                ).values_list("employee_id", flat=True))
            )

        if month: qs = qs.filter(month=int(month))
        if year:  qs = qs.filter(year=int(year))
        if emp_code: qs = qs.filter(employee_code__icontains=emp_code)
        if emp_name: qs = qs.filter(employee_name__icontains=emp_name)
        if dept and user.role == "ADMIN":
            qs = qs.filter(employee__department__iexact=dept)

        return Response([_rec(r) for r in qs.order_by("-date")[:500]])


class AttendanceSummaryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user     = request.user
        month    = request.query_params.get("month")
        year     = request.query_params.get("year")
        emp_code = request.query_params.get("employee_code")
        dept     = request.query_params.get("department")

        qs = AttendanceSummary.objects.select_related("employee").all()

        if user.role == "EMPLOYEE":
            try:
                emp = Employee.objects.get(user=user)
                qs = qs.filter(Q(employee=emp) | Q(employee_code=emp.employee_id))
            except Employee.DoesNotExist:
                return Response([])
        elif user.role == "MANAGER":
            dept_name = user.managed_department
            if not dept_name:
                return Response([])
            qs = qs.filter(employee__department__iexact=dept_name)

        if month: qs = qs.filter(month=int(month))
        if year:  qs = qs.filter(year=int(year))
        if emp_code: qs = qs.filter(employee_code__icontains=emp_code)
        if dept and user.role == "ADMIN":
            qs = qs.filter(employee__department__iexact=dept)

        return Response([_summ(s) for s in qs.order_by("employee_name")])


class MyAttendanceView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        now   = timezone.now()
        month = int(request.query_params.get("month", now.month))
        year  = int(request.query_params.get("year",  now.year))

        try:
            emp = Employee.objects.get(user=request.user)
        except Employee.DoesNotExist:
            return Response({"records": [], "summary": None})

        # ── Fast path: employee FK is already set (most common case after first load) ──
        records = EmployeeAttendance.objects.filter(
            employee=emp, month=month, year=year
        ).order_by("date")

        summary = AttendanceSummary.objects.filter(
            employee=emp, month=month, year=year
        ).first()

        # ── Slow path: FK not set yet — match by code/name, then auto-link ──
        if not records.exists():
            employee_codes = [c for c in (emp.employee_id, emp.employee_code, emp.user.username) if c]
            match_q = Q(employee_code__in=employee_codes)

            # Only add name matching if no code matched yet
            if not EmployeeAttendance.objects.filter(match_q, month=month, year=year).exists():
                full_name = emp.get_full_name().strip()
                if full_name:
                    match_q |= Q(employee_name__iexact=full_name)
                if emp.first_name and emp.last_name:
                    match_q |= (
                        Q(employee_name__icontains=emp.first_name) &
                        Q(employee_name__icontains=emp.last_name)
                    )

            records = EmployeeAttendance.objects.filter(
                match_q, month=month, year=year
            ).order_by("date")

            if not summary:
                summary = AttendanceSummary.objects.filter(
                    match_q, month=month, year=year
                ).first()

            # Auto-link in bulk — one query instead of N saves
            if records.exists():
                unlinked = [r for r in records if not r.employee_id]
                if unlinked:
                    for r in unlinked:
                        r.employee = emp
                    EmployeeAttendance.objects.bulk_update(unlinked, ['employee'])
                    # Refresh queryset to use the fast FK path next time
                    records = EmployeeAttendance.objects.filter(
                        employee=emp, month=month, year=year
                    ).order_by("date")

                if summary and not summary.employee_id:
                    summary.employee = emp
                    summary.save(update_fields=['employee'])

                if not emp.employee_code:
                    first_rec = records.first()
                    if first_rec and first_rec.employee_code:
                        emp.employee_code = first_rec.employee_code
                        emp.save(update_fields=['employee_code'])

        return Response({
            "records": [_rec(r) for r in records],
            "summary": _summ(summary) if summary else None,
        })


class AttendanceDashboardStatsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user  = request.user
        now   = timezone.now()
        month = int(request.query_params.get("month", now.month))
        year  = int(request.query_params.get("year",  now.year))

        qs = AttendanceSummary.objects.filter(month=month, year=year)

        if user.role == "MANAGER":
            dept_name = user.managed_department
            qs = qs.filter(
                employee__in=Employee.objects.filter(department__iexact=dept_name)
            ) if dept_name else qs.none()

        stats = qs.aggregate(
            avg_present=Avg("total_present"),
            avg_absent=Avg("total_absent"),
            total_late=Sum("total_late_mins"),
            total_present=Sum("total_present"),
            total_working=Sum("total_working_days"),
        )
        total = qs.count()
        avg_p = round(stats["avg_present"] or 0, 1)
        avg_a = round(stats["avg_absent"] or 0, 1)
        tot_lat = stats["total_late"] or 0
        total_present = stats["total_present"] or 0
        total_working = stats["total_working"] or 0
        avg_pct = round((total_present / total_working) * 100, 1) if total_working else 0

        top = list(qs.order_by("-total_absent").values(
            "employee_name", "employee_code", "total_absent", "total_present"
        )[:5])

        return Response({
            "month": month, "year": year,
            "total_employees": total,
            "avg_present_days": avg_p, "avg_absent_days": avg_a,
            "avg_attendance_pct": avg_pct, "total_late_minutes": tot_lat,
            "top_absentees": top,
        })


class IncompleteEmployeesView(APIView):
    """
    Returns employees that were auto-created from attendance uploads
    and still have incomplete profiles (no department, no email, etc.).
    Admin can use this to identify and complete employee records.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if request.user.role != "ADMIN":
            return Response({"error": "Admin only"}, status=403)

        # Auto-created employees have placeholder emails ending in @attendance.local
        # and empty department/phone fields
        incomplete = Employee.objects.filter(
            Q(email__endswith="@attendance.local") |
            Q(department="") |
            Q(phone="")
        ).select_related("user").order_by("first_name")

        data = []
        for emp in incomplete:
            missing = []
            if not emp.department:       missing.append("department")
            if not emp.designation:      missing.append("designation")
            if not emp.phone:            missing.append("phone")
            if emp.email.endswith("@attendance.local"): missing.append("email")
            if not emp.joining_date:     missing.append("joining date")

            data.append({
                "id": emp.id,
                "employee_id": emp.employee_id,
                "full_name": emp.get_full_name(),
                "first_name": emp.first_name,
                "last_name": emp.last_name,
                "department": emp.department or "",
                "designation": emp.designation or "",
                "phone": emp.phone or "",
                "email": emp.email if not emp.email.endswith("@attendance.local") else "",
                "joining_date": str(emp.joining_date) if emp.joining_date else "",
                "missing_fields": missing,
                "is_auto_created": emp.email.endswith("@attendance.local"),
                "username": emp.user.username,
            })

        return Response({
            "count": len(data),
            "employees": data,
        })


class AttendanceSyncView(APIView):
    """
    Re-links all EmployeeAttendance and AttendanceSummary records
    to their correct Employee objects after profile updates.
    Useful when an employee's employee_id or name changes.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if request.user.role != "ADMIN":
            return Response({"error": "Admin only"}, status=403)

        synced = 0
        for emp in Employee.objects.all():
            codes_to_match = []
            # Match by employee_code field (primary — the Excel code)
            if emp.employee_code:
                codes_to_match.append(emp.employee_code)
            # Match by employee_id (fallback)
            if emp.employee_id:
                codes_to_match.append(emp.employee_id)

            for code in codes_to_match:
                updated = EmployeeAttendance.objects.filter(
                    employee_code__iexact=code, employee__isnull=True
                ).update(employee=emp)
                synced += updated

                AttendanceSummary.objects.filter(
                    employee_code__iexact=code, employee__isnull=True
                ).update(employee=emp)

        return Response({
            "message": "Attendance records synced successfully.",
            "records_linked": synced,
        })


class AttendanceDateView(APIView):
    """
    Returns attendance records for a specific date, date range, or week.
    Supports role-based filtering (Admin=all, Manager=dept, Employee=own).

    Query params:
      date       YYYY-MM-DD  — single day
      start_date YYYY-MM-DD  — range start (use with end_date)
      end_date   YYYY-MM-DD  — range end
      week       YYYY-MM-DD  — any date in the week (Mon-Sun)
      employee_code          — filter by code (admin/manager only)
      employee_name          — filter by name (admin/manager only)
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from datetime import date as date_type, timedelta as td
        user = request.user

        # ── Resolve date range ────────────────────────────────────────────────
        date_str  = request.query_params.get("date")
        start_str = request.query_params.get("start_date")
        end_str   = request.query_params.get("end_date")
        week_str  = request.query_params.get("week")

        try:
            if week_str:
                anchor = date_type.fromisoformat(week_str)
                # Monday of that week
                start_dt = anchor - td(days=anchor.weekday())
                end_dt   = start_dt + td(days=6)
                mode = "week"
            elif start_str and end_str:
                start_dt = date_type.fromisoformat(start_str)
                end_dt   = date_type.fromisoformat(end_str)
                mode = "range"
            elif date_str:
                start_dt = end_dt = date_type.fromisoformat(date_str)
                mode = "day"
            else:
                # Default: today
                start_dt = end_dt = date_type.today()
                mode = "day"
        except ValueError:
            return Response({"error": "Invalid date format. Use YYYY-MM-DD."}, status=400)

        # ── Base queryset (existing rows) ─────────────────────────────────────
        qs = EmployeeAttendance.objects.select_related("employee").filter(
            date__gte=start_dt, date__lte=end_dt
        ).order_by("date", "employee_name")


        # ── Role-based filtering ──────────────────────────────────────────────
        if user.role == "EMPLOYEE":
            try:
                emp = Employee.objects.get(user=user)
                employee_codes = [code for code in (emp.employee_id, emp.employee_code) if code]
                qs = qs.filter(Q(employee=emp) | Q(employee_code__in=employee_codes))
            except Employee.DoesNotExist:
                return Response({"records": [], "summary": {}, "mode": mode})
        elif user.role == "MANAGER":
            dept_name = user.managed_department
            if not dept_name:
                return Response({"records": [], "summary": {}, "mode": mode})
            qs = qs.filter(
                Q(employee__department__iexact=dept_name) |
                Q(employee_code__in=Employee.objects.filter(
                    department__iexact=dept_name
                ).values_list("employee_id", flat=True))
            )

        # ── Optional extra filters (admin/manager) ────────────────────────────
        emp_code = request.query_params.get("employee_code")
        emp_name = request.query_params.get("employee_name")
        if emp_code: qs = qs.filter(employee_code__icontains=emp_code)
        if emp_name: qs = qs.filter(employee_name__icontains=emp_name)

        # ── Serialise ─────────────────────────────────────────────────────────
        records = []
        for r in qs:
            # Get department from linked employee or from designation field
            department = ""
            if r.employee and r.employee.department:
                department = r.employee.department
            
            # Get employee_id from linked employee or use employee_code
            employee_id = ""
            if r.employee and r.employee.employee_id:
                employee_id = r.employee.employee_id
            
            records.append({
                "id":            r.id,
                "date":          str(r.date),
                "day_name":      r.date.strftime("%A"),
                "employee_id":   employee_id,
                "employee_code": r.employee_code,
                "employee_name": r.employee_name,
                "department":    department,
                "designation":   r.designation,
                "in_time":       r.in_time.strftime("%H:%M")  if r.in_time  else None,
                "out_time":      r.out_time.strftime("%H:%M") if r.out_time else None,
                "working_hours": r.working_hours,
                "late_by":       r.late_by_display,
                "early_by":      r.early_by_display,
                "overtime":      _fmt_dur(r.overtime),
                "status":        r.status,
                "status_display": dict(EmployeeAttendance.STATUS_CHOICES).get(r.status, r.status),
                "is_present":    r.is_present,
                "is_absent":     r.is_absent,
                "is_weekly_off": r.is_weekly_off,
                "is_late":       r.is_late,
                "is_early_exit": r.is_early_exit,
                "is_missed_out": r.is_missed_out,
                "month":         r.month,
                "year":          r.year,
            })

        # ── Day-level summary ─────────────────────────────────────────────────
        total   = len(records)
        present = sum(1 for r in records if r["is_present"])
        absent  = sum(1 for r in records if r["is_absent"])
        wo      = sum(1 for r in records if r["is_weekly_off"])
        late    = sum(1 for r in records if r["is_late"])
        early   = sum(1 for r in records if r["is_early_exit"])

        summary = {
            "total":   total,
            "present": present,
            "absent":  absent,
            "weekly_off": wo,
            "late":    late,
            "early_exit": early,
            "attendance_pct": round(present / total * 100, 1) if total > 0 else 0,
            "date_range": {
                "start": str(start_dt),
                "end":   str(end_dt),
                "mode":  mode,
            },
        }

        return Response({"records": records, "summary": summary, "mode": mode})


def _fmt_dur(dur):
    """Format a timedelta as HH:MM string."""
    if not dur:
        return "00:00"
    total = int(dur.total_seconds())
    h, rem = divmod(total, 3600)
    m = rem // 60
    return "%02d:%02d" % (h, m)


class AttendanceRecordDetailView(APIView):
    """
    GET    /attendance/record/<id>/   — fetch one record
    PATCH  /attendance/record/<id>/   — edit in_time, out_time, status, remarks
    DELETE /attendance/record/<id>/   — delete one daily record (admin only)
    """
    permission_classes = [IsAuthenticated]

    def _get_record(self, pk, user):
        try:
            r = EmployeeAttendance.objects.get(pk=pk)
        except EmployeeAttendance.DoesNotExist:
            return None, Response({"error": "Record not found."}, status=404)

        # Employees can only see their own records
        if user.role == "EMPLOYEE":
            try:
                emp = Employee.objects.get(user=user)
                if r.employee != emp and r.employee_code != emp.employee_id:
                    return None, Response({"error": "Access denied."}, status=403)
            except Employee.DoesNotExist:
                return None, Response({"error": "Access denied."}, status=403)

        # Managers can only see their department
        elif user.role == "MANAGER":
            dept_name = user.managed_department
            if dept_name and r.employee:
                if r.employee.department.lower() != dept_name.lower():
                    return None, Response({"error": "Access denied."}, status=403)

        return r, None

    def get(self, request, pk):
        r, err = self._get_record(pk, request.user)
        if err:
            return err
        return Response(_rec(r))

    def patch(self, request, pk):
        if request.user.role not in ("ADMIN", "MANAGER"):
            return Response({"error": "Admin or Manager only."}, status=403)

        r, err = self._get_record(pk, request.user)
        if err:
            return err

        from datetime import time as time_type, timedelta

        def _parse_t(val):
            if not val:
                return None
            try:
                parts = str(val).strip().split(":")
                return time_type(int(parts[0]), int(parts[1]))
            except Exception:
                return None

        def _parse_d(val):
            if not val:
                return None
            try:
                parts = str(val).strip().split(":")
                h, m = int(parts[0]), int(parts[1])
                return timedelta(hours=h, minutes=m) if (h or m) else None
            except Exception:
                return None

        data = request.data

        if "in_time" in data:
            r.in_time = _parse_t(data["in_time"])
        if "out_time" in data:
            r.out_time = _parse_t(data["out_time"])
        if "status" in data:
            valid = [s[0] for s in EmployeeAttendance.STATUS_CHOICES]
            if data["status"] not in valid:
                return Response({"error": f"Invalid status. Choose from: {valid}"}, status=400)
            r.status = data["status"]
            r.is_present    = r.status in ("P", "P_MO", "HALF")
            r.is_absent     = r.status == "A"
            r.is_weekly_off = r.status == "WO"

        # Recalculate duration if both times are set
        if r.in_time and r.out_time:
            from datetime import datetime, date as date_cls
            dummy = date_cls(2000, 1, 1)
            dt_in  = datetime.combine(dummy, r.in_time)
            dt_out = datetime.combine(dummy, r.out_time)
            if dt_out > dt_in:
                r.duration = dt_out - dt_in
            # Late by: if in_time > 09:00
            shift_start = time_type(9, 0)
            if r.in_time > shift_start:
                dt_shift = datetime.combine(dummy, shift_start)
                r.late_by = dt_in - dt_shift
                r.is_late = True
            else:
                r.late_by = None
                r.is_late = False
            # Early by: if out_time < 18:00
            shift_end = time_type(18, 0)
            if r.out_time < shift_end:
                dt_shift_end = datetime.combine(dummy, shift_end)
                r.early_by = dt_shift_end - dt_out
                r.is_early_exit = True
            else:
                r.early_by = None
                r.is_early_exit = False

        r.save()

        # Rebuild monthly summary for this employee
        _build_summary(r.employee_code, r.employee_name, r.employee, r.month, r.year, r.upload)

        return Response({"message": "Record updated.", "record": _rec(r)})

    def delete(self, request, pk):
        if request.user.role != "ADMIN":
            return Response({"error": "Admin only."}, status=403)

        r, err = self._get_record(pk, request.user)
        if err:
            return err

        emp_code = r.employee_code
        emp_name = r.employee_name
        emp_obj  = r.employee
        month    = r.month
        year     = r.year
        upload   = r.upload
        r.delete()

        # Rebuild summary after deletion
        _build_summary(emp_code, emp_name, emp_obj, month, year, upload)

        return Response({"message": "Attendance record deleted."})
